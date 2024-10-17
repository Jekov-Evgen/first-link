import openpyxl
from graph import Graph
import numpy as np
from scipy.optimize import curve_fit

class Logics:
    def __init__(self, path: str) -> None:
        self.time = []
        self.value = [] 
        self.path = path 
        
    def read_data(self):
        data = openpyxl.open(self.path, read_only=True)
        sheet = data.active
        
        for row in range(1, sheet.max_row + 1):
            time_value = sheet.cell(row=row, column=1).value  
            data_value = sheet.cell(row=row, column=2).value 
            
            if time_value is not None and data_value is not None:
                self.time.append(time_value) 
                self.value.append(data_value)
        
    def draw_graph(self):
        Graph(self.time, self.value)
        
    def exp_function(self, t, K, T):
        return K * (1 - np.exp(-t / T))
    
    def calculations(self):
        time_array = np.array(self.time)
        value_array = np.array(self.value)
        
        initial_guess = [self.value[-1], 1] 
        
        popt, pcov = curve_fit(self.exp_function, time_array, value_array, p0=initial_guess)
        
        K_opt, T_opt = popt
        
        return [K_opt, T_opt]